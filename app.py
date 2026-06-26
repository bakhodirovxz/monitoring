import os, csv, io, threading
from datetime import datetime, timedelta, timezone
from typing import Optional, List

from fastapi import FastAPI, Request, Form, Depends, HTTPException
from fastapi.responses import HTMLResponse, RedirectResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from sqlalchemy.orm import Session

import models, auth, poller
from database import engine, SessionLocal, get_db

models.Base.metadata.create_all(bind=engine)

BASE_DIR  = os.path.dirname(os.path.abspath(__file__))
app       = FastAPI(title="HikCentral Monitor")
app.mount("/static", StaticFiles(directory=os.path.join(BASE_DIR, "static")), name="static")
templates = Jinja2Templates(directory=os.path.join(BASE_DIR, "templates"))


# ── AUTH HELPERS ───────────────────────────────────────────────────────
def _current_user(request: Request, db: Session = Depends(get_db)):
    return auth.get_current_user(db, request.cookies.get("access_token"))

def _branch_ids(user: models.User) -> Optional[List[int]]:
    if user.role == "superadmin":
        return None
    return [ub.branch_id for ub in user.user_branches]

def _nvr_ids(user: models.User, db) -> Optional[List[int]]:
    if user.role == "superadmin":
        return None
    nvr_ids = [un.nvr_id for un in user.user_nvrs]
    return nvr_ids if nvr_ids else None

def _cam_query(db: Session, user: models.User):
    ids = _branch_ids(user)
    nvr_ids = _nvr_ids(user, db)
    q = (db.query(models.Camera, models.NVR, models.Branch)
           .join(models.NVR, models.Camera.nvr_id == models.NVR.id, isouter=True)
           .join(models.Branch, models.NVR.branch_id == models.Branch.id, isouter=True))
    if nvr_ids is not None:
        q = q.filter(models.Camera.nvr_id.in_(nvr_ids))
    elif ids is not None:
        q = q.filter(models.NVR.branch_id.in_(ids))
    return q

def _event_query(db: Session, user: models.User):
    ids = _branch_ids(user)
    nvr_ids = _nvr_ids(user, db)
    q = (db.query(models.CameraEvent, models.Camera, models.NVR, models.Branch)
           .join(models.Camera, models.CameraEvent.camera_id == models.Camera.id)
           .join(models.NVR, models.Camera.nvr_id == models.NVR.id, isouter=True)
           .join(models.Branch, models.NVR.branch_id == models.Branch.id, isouter=True))
    if nvr_ids is not None:
        q = q.filter(models.Camera.nvr_id.in_(nvr_ids))
    elif ids is not None:
        q = q.filter(models.NVR.branch_id.in_(ids))
    return q


# ── PAGES ──────────────────────────────────────────────────────────────
@app.get("/", response_class=HTMLResponse)
async def index(request: Request, db: Session = Depends(get_db)):
    user = _current_user(request, db)
    return RedirectResponse("/dashboard" if user else "/login", status_code=302)


@app.get("/login", response_class=HTMLResponse)
async def login_page(request: Request):
    return templates.TemplateResponse("login.html", {"request": request, "error": None})


@app.post("/login", response_class=HTMLResponse)
async def login_post(request: Request, username: str = Form(...), password: str = Form(...),
                     db: Session = Depends(get_db)):
    user = db.query(models.User).filter(
        models.User.username == username, models.User.is_active == True).first()
    if not user or not auth.verify_password(password, user.password_hash):
        return templates.TemplateResponse("login.html",
                                          {"request": request, "error": "Login yoki parol xato"})
    token = auth.create_token({"sub": user.username})
    resp  = RedirectResponse("/dashboard", status_code=302)
    resp.set_cookie("access_token", token, httponly=True, max_age=3600 * 8)
    return resp


@app.get("/logout")
async def logout():
    resp = RedirectResponse("/login", status_code=302)
    resp.delete_cookie("access_token")
    return resp


@app.get("/dashboard", response_class=HTMLResponse)
async def dashboard(request: Request, db: Session = Depends(get_db)):
    user = _current_user(request, db)
    if not user:
        return RedirectResponse("/login", status_code=302)
    ids      = _branch_ids(user)
    bq       = db.query(models.Branch)
    branches = bq.filter(models.Branch.id.in_(ids)).all() if ids is not None else bq.all()
    return templates.TemplateResponse("dashboard.html",
                                      {"request": request, "user": user, "branches": branches})


@app.get("/reports", response_class=HTMLResponse)
async def reports_page(request: Request, db: Session = Depends(get_db)):
    user = _current_user(request, db)
    if not user:
        return RedirectResponse("/login", status_code=302)
    ids      = _branch_ids(user)
    bq       = db.query(models.Branch)
    branches = bq.filter(models.Branch.id.in_(ids)).all() if ids is not None else bq.all()
    return templates.TemplateResponse("reports.html",
                                      {"request": request, "user": user, "branches": branches})


@app.get("/manage", response_class=HTMLResponse)
async def manage_page(request: Request, db: Session = Depends(get_db)):
    user = _current_user(request, db)
    if not user:
        return RedirectResponse("/login", status_code=302)
    ids = _branch_ids(user)
    if ids is not None:
        branches = db.query(models.Branch).filter(models.Branch.id.in_(ids)).all() if ids else []
        nvrs = db.query(models.NVR).filter(models.NVR.branch_id.in_(ids)).order_by(models.NVR.name).all() if ids else []
        managed_users = (db.query(models.User)
            .join(models.UserBranch)
            .filter(models.UserBranch.branch_id.in_(ids),
                    models.User.role != "superadmin")
            .distinct().all()) if ids else []
    else:
        branches = db.query(models.Branch).all()
        nvrs = db.query(models.NVR).order_by(models.NVR.name).all()
        managed_users = []
    return templates.TemplateResponse("manage.html", {
        "request": request, "user": user,
        "branches": branches, "nvrs": nvrs, "managed_users": managed_users,
    })


@app.post("/api/manage/users")
async def create_managed_user(request: Request, db: Session = Depends(get_db)):
    cur = _current_user(request, db)
    if not cur:
        raise HTTPException(status_code=401)
    ids = _branch_ids(cur)
    if not ids:
        raise HTTPException(status_code=403, detail="Faqat filial adminlari foydalanuvchi yarata oladi")
    data = await request.json()
    if not data.get("username") or not data.get("password"):
        raise HTTPException(status_code=400, detail="Login va parol kerak")
    if db.query(models.User).filter(models.User.username == data["username"]).first():
        raise HTTPException(status_code=400, detail="Bu login allaqachon mavjud")
    branch_ids = [bid for bid in data.get("branch_ids", []) if bid in ids]
    if not branch_ids:
        branch_ids = ids[:1]
    u = models.User(
        username=data["username"],
        password_hash=auth.hash_password(data["password"]),
        full_name=data.get("full_name", ""),
        role="branch_admin",
    )
    db.add(u); db.flush()
    for bid in branch_ids:
        db.add(models.UserBranch(user_id=u.id, branch_id=bid))
    for nid in data.get("nvr_ids", []):
        nvr = db.query(models.NVR).get(nid)
        if nvr and nvr.branch_id in ids:
            db.add(models.UserNVR(user_id=u.id, nvr_id=nid))
    db.commit()
    return {"id": u.id, "username": u.username}


@app.put("/api/manage/users/{uid}")
async def update_managed_user(uid: int, request: Request, db: Session = Depends(get_db)):
    cur = _current_user(request, db)
    if not cur:
        raise HTTPException(status_code=401)
    cur_ids = _branch_ids(cur)
    if not cur_ids:
        raise HTTPException(status_code=403)
    u = db.query(models.User).get(uid)
    if not u:
        raise HTTPException(status_code=404)
    user_bids = [ub.branch_id for ub in u.user_branches]
    if not any(bid in cur_ids for bid in user_bids):
        raise HTTPException(status_code=403)
    data = await request.json()
    for field in ("full_name", "is_active"):
        if field in data:
            setattr(u, field, data[field])
    if data.get("password"):
        u.password_hash = auth.hash_password(data["password"])
    if "branch_ids" in data:
        db.query(models.UserBranch).filter(models.UserBranch.user_id == uid).delete()
        for bid in data["branch_ids"]:
            if bid in cur_ids and db.query(models.Branch).get(bid):
                db.add(models.UserBranch(user_id=uid, branch_id=bid))
    if "nvr_ids" in data:
        db.query(models.UserNVR).filter(models.UserNVR.user_id == uid).delete()
        for nid in data["nvr_ids"]:
            nvr = db.query(models.NVR).get(nid)
            if nvr and nvr.branch_id in cur_ids:
                db.add(models.UserNVR(user_id=uid, nvr_id=nid))
    db.commit()
    return {"ok": True}


@app.delete("/api/manage/users/{uid}")
async def delete_managed_user(uid: int, request: Request, db: Session = Depends(get_db)):
    cur = _current_user(request, db)
    if not cur:
        raise HTTPException(status_code=401)
    if cur.id == uid:
        raise HTTPException(status_code=400, detail="O'zingizni o'chira olmaysiz")
    cur_ids = _branch_ids(cur)
    if not cur_ids:
        raise HTTPException(status_code=403)
    u = db.query(models.User).get(uid)
    if not u:
        raise HTTPException(status_code=404)
    user_bids = [ub.branch_id for ub in u.user_branches]
    if not any(bid in cur_ids for bid in user_bids):
        raise HTTPException(status_code=403)
    if u.role == "superadmin":
        raise HTTPException(status_code=400, detail="Superadminni o'chirish mumkin emas")
    db.delete(u); db.commit()
    return {"ok": True}


@app.get("/admin", response_class=HTMLResponse)
async def admin_page(request: Request, db: Session = Depends(get_db)):
    user = _current_user(request, db)
    if not user:
        return RedirectResponse("/login", status_code=302)
    if user.role != "superadmin":
        return RedirectResponse("/dashboard", status_code=302)
    branches = db.query(models.Branch).all()
    users    = db.query(models.User).all()
    nvrs     = db.query(models.NVR).all()
    return templates.TemplateResponse("admin.html", {
        "request": request, "user": user,
        "branches": branches, "users": users, "nvrs": nvrs,
    })


# ── API: STATUS ────────────────────────────────────────────────────────
@app.get("/api/cameras")
async def api_cameras(request: Request,
                       branch_id: Optional[int] = None,
                       status: Optional[int] = None,
                       db: Session = Depends(get_db)):
    user = _current_user(request, db)
    if not user:
        raise HTTPException(status_code=401)
    q = _cam_query(db, user)
    if branch_id:
        q = q.filter(models.NVR.branch_id == branch_id)
    if status is not None:
        q = q.filter(models.Camera.current_status == status)
    rows = q.order_by(models.Camera.name).all()
    return {"cameras": [_cam_row(c, n, b) for c, n, b in rows]}


LOCAL_TZ = timezone(timedelta(hours=5))

def _cam_row(cam, nvr, branch):
    os = None
    if cam.current_status == 2:
        ts = poller.offline_since.get(cam.hik_code)
        if ts:
            os = ts.astimezone(LOCAL_TZ).replace(tzinfo=None).isoformat()
    return {
        "id":            cam.id,
        "name":          cam.name,
        "ip":            cam.channel_ip,
        "status":        cam.current_status,
        "offline_since": os,
        "nvr_id":        nvr.id if nvr else None,
        "nvr_name":      nvr.name if nvr else "",
        "nvr_ip":        nvr.ip if nvr else "",
        "branch":        branch.name if branch else "Tayinlanmagan",
        "branch_id":     branch.id if branch else None,
    }


@app.get("/api/stats")
async def api_stats(request: Request, db: Session = Depends(get_db)):
    user = _current_user(request, db)
    if not user:
        raise HTTPException(status_code=401)
    rows = _cam_query(db, user).all()
    cams = [c for c, _, _ in rows]
    return {
        "total":       len(cams),
        "online":      sum(1 for c in cams if c.current_status == 1),
        "offline":     sum(1 for c in cams if c.current_status == 2),
        "unknown":     sum(1 for c in cams if c.current_status == 0),
        "last_update": _local_iso(poller.last_poll_time) if poller.last_poll_time else None,
    }


# ── API: EVENTS ────────────────────────────────────────────────────────
@app.get("/api/events")
async def api_events(request: Request,
                      branch_id: Optional[int] = None,
                      nvr_id: Optional[int] = None,
                      camera_id: Optional[int] = None,
                      event_type: Optional[str] = None,
                      search: Optional[str] = None,
                      date_from: Optional[str] = None,
                      date_to: Optional[str] = None,
                      limit: int = 100,
                      offset: int = 0,
                      db: Session = Depends(get_db)):
    user = _current_user(request, db)
    if not user:
        raise HTTPException(status_code=401)
    q = _event_query(db, user)
    if branch_id:
        q = q.filter(models.NVR.branch_id == branch_id)
    if nvr_id:
        q = q.filter(models.NVR.id == nvr_id)
    if camera_id:
        q = q.filter(models.CameraEvent.camera_id == camera_id)
    if event_type:
        q = q.filter(models.CameraEvent.event_type == event_type)
    if search:
        q = q.filter(models.Camera.name.ilike(f"%{search}%"))
    q = _date_filter_utc(q, models.CameraEvent.started_at, date_from, date_to)

    # Aggregate stats across ALL matching rows (not just current page)
    from sqlalchemy import func
    total       = q.count()
    total_off   = q.filter(models.CameraEvent.event_type == "offline").count()
    total_on    = total - total_off
    total_dur   = (q.filter(models.CameraEvent.event_type == "offline",
                            models.CameraEvent.duration_sec.isnot(None))
                    .with_entities(func.sum(models.CameraEvent.duration_sec)).scalar()) or 0
    avg_dur     = (q.filter(models.CameraEvent.event_type == "offline",
                            models.CameraEvent.duration_sec.isnot(None))
                    .with_entities(func.avg(models.CameraEvent.duration_sec)).scalar())
    unique_cams = q.with_entities(models.Camera.id).distinct().count()

    rows = q.order_by(models.CameraEvent.started_at.desc()).offset(offset).limit(limit).all()
    return {
        "total": total,
        "total_offline": total_off,
        "total_online": total_on,
        "total_duration_sec": total_dur,
        "avg_duration_sec": avg_dur,
        "unique_cameras": unique_cams,
        "events": [_ev_row(ev, c, n, b) for ev, c, n, b in rows],
    }


def _local_iso(dt):
    if dt is None:
        return None
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    return dt.astimezone(LOCAL_TZ).replace(tzinfo=None).isoformat()

def _local_str(dt):
    if dt is None:
        return "-"
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    return dt.astimezone(LOCAL_TZ).strftime("%d.%m.%Y %H:%M:%S")

def _date_filter_utc(q, col, date_from, date_to):
    """Convert local date boundaries to UTC before filtering."""
    offset = timedelta(hours=5)
    if date_from:
        try:
            local_start = datetime.fromisoformat(date_from)
            q = q.filter(col >= local_start - offset)
        except Exception:
            pass
    if date_to:
        try:
            local_end = datetime.fromisoformat(date_to) + timedelta(days=1)
            q = q.filter(col < local_end - offset)
        except Exception:
            pass
    return q

def _ev_row(ev, cam, nvr, branch):
    return {
        "id":          ev.id,
        "camera_id":   cam.id,
        "camera_name": cam.name,
        "camera_ip":   cam.channel_ip,
        "branch":      branch.name if branch else "Tayinlanmagan",
        "nvr_name":    nvr.name if nvr else "",
        "event_type":  ev.event_type,
        "started_at":  _local_iso(ev.started_at),
        "ended_at":    _local_iso(ev.ended_at),
        "duration_sec": ev.duration_sec,
    }


# ── API: ADMIN ─────────────────────────────────────────────────────────
@app.post("/api/branches")
async def create_branch(request: Request, db: Session = Depends(get_db)):
    user = _current_user(request, db)
    if not user or user.role != "superadmin":
        raise HTTPException(status_code=403)
    data = await request.json()
    name = data.get("name", "").strip()
    if not name:
        raise HTTPException(status_code=400, detail="Nom kerak")
    if db.query(models.Branch).filter(models.Branch.name == name).first():
        raise HTTPException(status_code=400, detail="Bu nom allaqachon mavjud")
    b = models.Branch(name=name)
    db.add(b); db.commit(); db.refresh(b)
    return {"id": b.id, "name": b.name}


@app.delete("/api/branches/{bid}")
async def delete_branch(bid: int, request: Request, db: Session = Depends(get_db)):
    user = _current_user(request, db)
    if not user or user.role != "superadmin":
        raise HTTPException(status_code=403)
    b = db.query(models.Branch).get(bid)
    if not b:
        raise HTTPException(status_code=404)
    db.delete(b); db.commit()
    return {"ok": True}


@app.post("/api/users")
async def create_user(request: Request, db: Session = Depends(get_db)):
    user = _current_user(request, db)
    if not user or user.role != "superadmin":
        raise HTTPException(status_code=403)
    data = await request.json()
    if not data.get("username") or not data.get("password"):
        raise HTTPException(status_code=400, detail="Login va parol kerak")
    if db.query(models.User).filter(models.User.username == data["username"]).first():
        raise HTTPException(status_code=400, detail="Bu login allaqachon mavjud")
    u = models.User(
        username      = data["username"],
        password_hash = auth.hash_password(data["password"]),
        full_name     = data.get("full_name", ""),
        role          = data.get("role", "branch_admin"),
    )
    db.add(u); db.flush()
    for bid in data.get("branch_ids", []):
        if db.query(models.Branch).get(bid):
            db.add(models.UserBranch(user_id=u.id, branch_id=bid))
    db.commit()
    return {"id": u.id, "username": u.username}


@app.put("/api/users/{uid}")
async def update_user(uid: int, request: Request, db: Session = Depends(get_db)):
    cur = _current_user(request, db)
    if not cur or cur.role != "superadmin":
        raise HTTPException(status_code=403)
    data = await request.json()
    u    = db.query(models.User).get(uid)
    if not u:
        raise HTTPException(status_code=404)
    for field in ("full_name", "role", "is_active"):
        if field in data:
            setattr(u, field, data[field])
    if data.get("password"):
        u.password_hash = auth.hash_password(data["password"])
    if "branch_ids" in data:
        db.query(models.UserBranch).filter(models.UserBranch.user_id == u.id).delete()
        for bid in data["branch_ids"]:
            if db.query(models.Branch).get(bid):
                db.add(models.UserBranch(user_id=u.id, branch_id=bid))
    db.commit()
    return {"ok": True}


@app.delete("/api/users/{uid}")
async def delete_user(uid: int, request: Request, db: Session = Depends(get_db)):
    cur = _current_user(request, db)
    if not cur or cur.role != "superadmin":
        raise HTTPException(status_code=403)
    if cur.id == uid:
        raise HTTPException(status_code=400, detail="O'zingizni o'chira olmaysiz")
    u = db.query(models.User).get(uid)
    if not u:
        raise HTTPException(status_code=404)
    if u.role == "superadmin":
        raise HTTPException(status_code=400, detail="Superadminni o'chirish mumkin emas")
    db.delete(u); db.commit()
    return {"ok": True}


@app.post("/api/nvrs/{nid}/assign")
async def assign_nvr(nid: int, request: Request, db: Session = Depends(get_db)):
    cur = _current_user(request, db)
    if not cur or cur.role != "superadmin":
        raise HTTPException(status_code=403)
    data = await request.json()
    nvr  = db.query(models.NVR).get(nid)
    if not nvr:
        raise HTTPException(status_code=404)
    nvr.branch_id = data.get("branch_id")
    db.commit()
    return {"ok": True}


@app.post("/api/sync-nvrs")
async def sync_nvrs(request: Request, db: Session = Depends(get_db)):
    cur = _current_user(request, db)
    if not cur or cur.role != "superadmin":
        raise HTTPException(status_code=403)
    poller.sync_nvrs_to_db(db)
    return {"ok": True, "count": db.query(models.NVR).count()}


# ── CAMERA PROFILE ─────────────────────────────────────────────────────
@app.get("/camera/{cam_id}", response_class=HTMLResponse)
async def camera_profile(cam_id: int, request: Request, db: Session = Depends(get_db)):
    user = _current_user(request, db)
    if not user:
        return RedirectResponse("/login", status_code=302)
    ids = _branch_ids(user)
    q = _cam_query(db, user).filter(models.Camera.id == cam_id)
    row = q.first()
    if not row:
        raise HTTPException(status_code=404)
    cam, nvr, branch = row
    branches = db.query(models.Branch)
    if ids is not None:
        branches = branches.filter(models.Branch.id.in_(ids))
    return templates.TemplateResponse("camera_detail.html", {
        "request": request, "user": user,
        "camera": cam, "nvr": nvr, "branch": branch,
        "branches": branches.all(),
    })


@app.get("/api/cameras/{cam_id}")
async def api_camera_detail(cam_id: int, request: Request, db: Session = Depends(get_db)):
    user = _current_user(request, db)
    if not user:
        raise HTTPException(status_code=401)
    q = _cam_query(db, user).filter(models.Camera.id == cam_id)
    row = q.first()
    if not row:
        raise HTTPException(status_code=404)
    cam, nvr, branch = row
    offline_count = db.query(models.CameraEvent).filter(
        models.CameraEvent.camera_id == cam.id,
        models.CameraEvent.event_type == "offline",
    ).count()
    last_events = (db.query(models.CameraEvent)
                   .filter(models.CameraEvent.camera_id == cam.id)
                   .order_by(models.CameraEvent.started_at.desc())
                   .limit(10).all())
    return {
        **_cam_row(cam, nvr, branch),
        "hik_code": cam.hik_code,
        "last_status_change": _local_iso(cam.last_status_change),
        "offline_count": offline_count,
        "last_events": [_ev_row(ev, cam, nvr, branch) for ev in last_events],
    }


@app.get("/api/cameras/{cam_id}/daily-stats")
async def api_camera_daily_stats(cam_id: int, request: Request, db: Session = Depends(get_db)):
    user = _current_user(request, db)
    if not user:
        raise HTTPException(status_code=401)
    q = _cam_query(db, user).filter(models.Camera.id == cam_id)
    row = q.first()
    if not row:
        raise HTTPException(status_code=404)
    cam = row[0]
    from sqlalchemy import func, cast, Date
    rows = (db.query(
                cast(models.CameraEvent.started_at, Date).label("day"),
                func.sum(models.CameraEvent.duration_sec).label("total_sec"),
                func.count().label("count"))
            .filter(models.CameraEvent.camera_id == cam.id,
                    models.CameraEvent.event_type == "offline")
            .group_by("day")
            .order_by("day")
            .all())
    return {
        "daily": [
            {"day": str(r.day), "total_sec": r.total_sec or 0, "count": r.count}
            for r in rows
        ]
    }


@app.put("/api/cameras/{cam_id}")
async def update_camera(cam_id: int, request: Request, db: Session = Depends(get_db)):
    user = _current_user(request, db)
    if not user:
        raise HTTPException(status_code=401)
    cam = db.query(models.Camera).filter(models.Camera.id == cam_id).first()
    if not cam:
        raise HTTPException(status_code=404)
    ids = _branch_ids(user)
    if ids is not None:
        nvr = db.query(models.NVR).filter(models.NVR.id == cam.nvr_id).first()
        if not nvr or nvr.branch_id not in ids:
            raise HTTPException(status_code=403)
    data = await request.json()
    if "name" in data:
        cam.name = data["name"]
        cam.name_overridden = True
        db.commit()
    return {"ok": True}


@app.put("/api/nvrs/{nvr_id}")
async def update_nvr(nvr_id: int, request: Request, db: Session = Depends(get_db)):
    user = _current_user(request, db)
    if not user:
        raise HTTPException(status_code=401)
    nvr = db.query(models.NVR).filter(models.NVR.id == nvr_id).first()
    if not nvr:
        raise HTTPException(status_code=404)
    ids = _branch_ids(user)
    if ids is not None and nvr.branch_id not in ids:
        raise HTTPException(status_code=403)
    data = await request.json()
    if "name" in data:
        nvr.name = data["name"]
        nvr.name_overridden = True
        db.commit()
    return {"ok": True}


@app.get("/api/branches/{bid}/nvrs")
async def api_branch_nvrs(bid: int, request: Request, db: Session = Depends(get_db)):
    user = _current_user(request, db)
    if not user:
        raise HTTPException(status_code=401)
    nvrs = db.query(models.NVR).filter(models.NVR.branch_id == bid).order_by(models.NVR.name).all()
    return {"nvrs": [{"id": n.id, "name": n.name or n.hik_code, "ip": n.ip} for n in nvrs]}


# ── SNAPSHOT ───────────────────────────────────────────────────────────
import base64 as _b64
SNAP_DIR = os.path.join(BASE_DIR, "static", "snapshots")
os.makedirs(SNAP_DIR, exist_ok=True)

@app.get("/api/cameras/{cam_id}/snapshot")
async def camera_snapshot(cam_id: int, request: Request, db: Session = Depends(get_db)):
    """Return the stored snapshot file (no HikCentral call)."""
    user = _current_user(request, db)
    if not user:
        raise HTTPException(status_code=401)
    cam = db.query(models.Camera).filter(models.Camera.id == cam_id).first()
    if not cam:
        raise HTTPException(status_code=404)
    path = os.path.join(SNAP_DIR, f"{cam_id}.jpg")
    if os.path.exists(path):
        return {"has_snapshot": True, "url": f"/static/snapshots/{cam_id}.jpg?v={int(os.path.getmtime(path))}"}
    return {"has_snapshot": False, "url": None}


@app.post("/api/cameras/{cam_id}/snapshot/refresh")
async def camera_snapshot_refresh(cam_id: int, request: Request, db: Session = Depends(get_db)):
    """Capture a fresh snapshot from HikCentral and save to disk."""
    user = _current_user(request, db)
    if not user:
        raise HTTPException(status_code=401)
    cam = db.query(models.Camera).filter(models.Camera.id == cam_id).first()
    if not cam:
        raise HTTPException(status_code=404)
    result = poller.hik("/artemis/api/video/v1/camera/capture",
                        {"cameraIndexCode": cam.hik_code})
    if not result or str(result.get("code")) != "0":
        raise HTTPException(status_code=502, detail="HikCentral dan rasm olinmadi")
    data = result.get("data", "")
    if not data:
        raise HTTPException(status_code=502, detail="Bo'sh rasm")
    if "," in data:
        data = data.split(",", 1)[1]
    img_bytes = _b64.b64decode(data)
    path = os.path.join(SNAP_DIR, f"{cam_id}.jpg")
    with open(path, "wb") as f:
        f.write(img_bytes)
    return {"ok": True, "url": f"/static/snapshots/{cam_id}.jpg?v={int(os.path.getmtime(path))}"}


# ── API: EXPORT ────────────────────────────────────────────────────────
def _export_rows(db, user, branch_id, date_from, date_to, nvr_id=None):
    q = _event_query(db, user)
    if branch_id:
        try:
            q = q.filter(models.NVR.branch_id == int(branch_id))
        except Exception:
            pass
    if nvr_id:
        try:
            q = q.filter(models.NVR.id == int(nvr_id))
        except Exception:
            pass
    q = _date_filter_utc(q, models.CameraEvent.started_at, date_from, date_to)
    return q.order_by(models.CameraEvent.started_at.desc()).all()


@app.get("/api/export/csv")
async def export_csv(request: Request,
                      branch_id: Optional[str] = None,
                      nvr_id: Optional[str] = None,
                      date_from: Optional[str] = None,
                      date_to: Optional[str] = None,
                      db: Session = Depends(get_db)):
    user = _current_user(request, db)
    if not user:
        raise HTTPException(status_code=401)
    rows   = _export_rows(db, user, branch_id, date_from, date_to, nvr_id)
    output = io.StringIO()
    w      = csv.writer(output)
    w.writerow(["Kamera", "IP", "Filial", "NVR", "Hodisa", "Boshlanish", "Tugash", "Davomiyligi (daq.)"])
    for ev, cam, nvr, branch in rows:
        dur = f"{ev.duration_sec/60:.1f}" if ev.duration_sec else "-"
        w.writerow([
            cam.name, cam.channel_ip,
            branch.name if branch else "-",
            nvr.name if nvr else "-",
            "Offline" if ev.event_type == "offline" else "Online",
            _local_str(ev.started_at),
            _local_str(ev.ended_at),
            dur,
        ])
    fname = f"hisobot_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
    return StreamingResponse(
        io.BytesIO(output.getvalue().encode("utf-8-sig")),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={fname}"},
    )


@app.get("/api/export/excel")
async def export_excel(request: Request,
                        branch_id: Optional[str] = None,
                        nvr_id: Optional[str] = None,
                        date_from: Optional[str] = None,
                        date_to: Optional[str] = None,
                        db: Session = Depends(get_db)):
    user = _current_user(request, db)
    if not user:
        raise HTTPException(status_code=401)
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise HTTPException(status_code=500, detail="pip install openpyxl")

    rows = _export_rows(db, user, branch_id, date_from, date_to, nvr_id)
    wb   = Workbook()
    ws   = wb.active
    ws.title = "Hisobot"
    headers  = ["Kamera", "IP", "Filial", "NVR", "Hodisa", "Boshlanish", "Tugash", "Davomiyligi (daq.)"]
    hfont    = Font(bold=True, color="FFFFFF")
    hfill    = PatternFill("solid", fgColor="1a5276")
    for col, h in enumerate(headers, 1):
        cell            = ws.cell(row=1, column=col, value=h)
        cell.font       = hfont
        cell.fill       = hfill
        cell.alignment  = Alignment(horizontal="center")
    for ev, cam, nvr, branch in rows:
        dur = round(ev.duration_sec / 60, 1) if ev.duration_sec else None
        ws.append([
            cam.name, cam.channel_ip,
            branch.name if branch else "-",
            nvr.name if nvr else "-",
            "Offline" if ev.event_type == "offline" else "Online",
            _local_str(ev.started_at),
            _local_str(ev.ended_at),
            dur,
        ])
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = min(
            max(len(str(cell.value or "")) for cell in col) + 4, 40)
    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    fname = f"hisobot_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"},
    )


# ── STARTUP ────────────────────────────────────────────────────────────
@app.on_event("startup")
async def on_startup():
    db = SessionLocal()
    try:
        if db.query(models.User).count() == 0:
            db.add(models.User(
                username      = "admin",
                password_hash = auth.hash_password("admin123"),
                full_name     = "Super Admin",
                role          = "superadmin",
            ))
            db.commit()
    finally:
        db.close()

    poller.set_session_factory(SessionLocal)
    t = threading.Thread(target=poller.run_forever, daemon=True)
    t.start()
